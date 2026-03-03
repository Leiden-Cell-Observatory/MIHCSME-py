"""Tests for write_metadata_to_excel."""

import io

import openpyxl
import pytest

from mihcsme_py.models import (
    AssayCondition,
    DataOwner,
    InvestigationInfo,
    InvestigationInformation,
    MIHCSMEMetadata,
)
from mihcsme_py.writer import DEFAULT_CONDITION_KEYS, write_metadata_to_excel


def _write(metadata: MIHCSMEMetadata) -> openpyxl.Workbook:
    buf = io.BytesIO()
    write_metadata_to_excel(metadata, buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf)


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------


def test_empty_metadata_writes_all_sheets():
    wb = _write(MIHCSMEMetadata())
    assert set(wb.sheetnames) == {
        "InvestigationInformation",
        "StudyInformation",
        "AssayInformation",
        "AssayConditions",
    }


def test_partially_populated_metadata_writes_all_sheets():
    meta = MIHCSMEMetadata(
        investigation_information=InvestigationInformation(
            data_owner=DataOwner(first_name="Jane", last_name="Doe")
        )
    )
    wb = _write(meta)
    assert "StudyInformation" in wb.sheetnames
    assert "AssayInformation" in wb.sheetnames
    assert "AssayConditions" in wb.sheetnames


# ---------------------------------------------------------------------------
# AssayConditions – empty case uses default headers
# ---------------------------------------------------------------------------


def test_assay_conditions_empty_writes_default_headers():
    wb = _write(MIHCSMEMetadata())
    ws = wb["AssayConditions"]
    header_row = [ws.cell(row=2, column=c).value for c in range(1, 2 + len(DEFAULT_CONDITION_KEYS) + 1)]
    assert header_row[0] == "Plate"
    assert header_row[1] == "Well"
    assert header_row[2:] == DEFAULT_CONDITION_KEYS


def test_assay_conditions_empty_has_no_data_rows():
    wb = _write(MIHCSMEMetadata())
    ws = wb["AssayConditions"]
    # Row 3 onwards should be empty
    assert ws.cell(row=3, column=1).value is None


# ---------------------------------------------------------------------------
# AssayConditions – populated case
# ---------------------------------------------------------------------------


def test_assay_conditions_populated_writes_data():
    meta = MIHCSMEMetadata(
        assay_conditions=[
            AssayCondition(plate="Plate1", well="A01", conditions={"Treatment": "DMSO", "Dose": "0"}),
            AssayCondition(plate="Plate1", well="B02", conditions={"Treatment": "DrugX", "Dose": "10"}),
        ]
    )
    wb = _write(meta)
    ws = wb["AssayConditions"]

    # Headers in row 2
    headers = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert "Plate" in headers
    assert "Well" in headers
    assert "Dose" in headers
    assert "Treatment" in headers

    # First data row
    assert ws.cell(row=3, column=1).value == "Plate1"
    assert ws.cell(row=3, column=2).value == "A01"


def test_assay_conditions_column_order_is_sorted():
    meta = MIHCSMEMetadata(
        assay_conditions=[
            AssayCondition(plate="P1", well="A01", conditions={"Zebra": "z", "Apple": "a"}),
        ]
    )
    wb = _write(meta)
    ws = wb["AssayConditions"]
    col3 = ws.cell(row=2, column=3).value
    col4 = ws.cell(row=2, column=4).value
    assert col3 == "Apple"
    assert col4 == "Zebra"


# ---------------------------------------------------------------------------
# Grouped sheets – Investigation data round-trip
# ---------------------------------------------------------------------------


def test_investigation_information_values_written():
    meta = MIHCSMEMetadata(
        investigation_information=InvestigationInformation(
            data_owner=DataOwner(first_name="Jane", last_name="Doe"),
            investigation_info=InvestigationInfo(investigation_title="My Study"),
        )
    )
    wb = _write(meta)
    ws = wb["InvestigationInformation"]

    values = {ws.cell(row=r, column=2).value: ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)}
    assert values.get("First Name") == "Jane"
    assert values.get("Last Name") == "Doe"
    assert values.get("Investigation Title") == "My Study"


def test_empty_grouped_sheet_has_column_headers():
    wb = _write(MIHCSMEMetadata())
    for sheet_name in ("InvestigationInformation", "StudyInformation", "AssayInformation"):
        ws = wb[sheet_name]
        # The header row (row 2 when a comment is present) must have Group/Key/Value
        headers = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Group" in headers, f"{sheet_name} missing 'Group' header"
