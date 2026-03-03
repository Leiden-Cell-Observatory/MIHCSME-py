"""Write MIHCSME metadata to Excel format."""

from pathlib import Path
from typing import Dict, List, Any, Union, BinaryIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from mihcsme_py.models import MIHCSMEMetadata

# Default condition column headers written when AssayConditions sheet is empty
DEFAULT_CONDITION_KEYS = ["Treatment", "Concentration", "Unit", "CellLine", "TimeTreatment", "RepID"]


def write_metadata_to_excel(
    metadata: MIHCSMEMetadata, output_path: Union[Path, BinaryIO]
) -> None:
    """
    Write MIHCSME metadata to Excel file.

    All sheets are always written, even if empty, so the file can be used as
    a template. The AssayConditions sheet uses default column headers when no
    conditions are present.

    :param metadata: MIHCSMEMetadata object to export
    :param output_path: Path to output Excel file, or a file-like object (e.g., BytesIO)
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Write Investigation Information (always)
    groups = metadata.investigation_information.groups if metadata.investigation_information else {}
    _write_grouped_sheet(
        wb,
        "InvestigationInformation",
        groups,
        header_comment="# Investigation Information - Metadata about the overall investigation"
    )

    # Write Study Information (always)
    groups = metadata.study_information.groups if metadata.study_information else {}
    _write_grouped_sheet(
        wb,
        "StudyInformation",
        groups,
        header_comment="# Study Information - Metadata about the study design"
    )

    # Write Assay Information (always)
    groups = metadata.assay_information.groups if metadata.assay_information else {}
    _write_grouped_sheet(
        wb,
        "AssayInformation",
        groups,
        header_comment="# Assay Information - Metadata about the assay protocol"
    )

    # Write Assay Conditions (always)
    _write_assay_conditions(wb, metadata.assay_conditions)

    # Write Reference Sheets
    for ref_sheet in metadata.reference_sheets:
        _write_reference_sheet(wb, ref_sheet.name, ref_sheet.data)

    # Save workbook
    wb.save(output_path)


def _write_grouped_sheet(
    wb: Workbook,
    sheet_name: str,
    groups: Dict[str, Dict[str, str]],
    header_comment: str = None
) -> None:
    """
    Write a grouped metadata sheet (Investigation/Study/Assay Information).

    :param wb: Workbook object
    :param sheet_name: Name of the sheet
    :param groups: Dictionary of groups {group_name: {key: value}}
    :param header_comment: Optional comment to add at the top
    """
    ws = wb.create_sheet(sheet_name)

    # Add header comment if provided
    row_num = 1
    if header_comment:
        ws.cell(row=row_num, column=1, value=header_comment)
        ws.cell(row=row_num, column=1).font = Font(italic=True, color="808080")
        row_num += 1

    # Add column headers
    ws.cell(row=row_num, column=1, value="Group")
    ws.cell(row=row_num, column=2, value="Key")
    ws.cell(row=row_num, column=3, value="Value")

    # Style headers
    for col in range(1, 4):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row_num += 1

    # Write data
    for group_name, group_data in groups.items():
        for key, value in group_data.items():
            ws.cell(row=row_num, column=1, value=group_name)
            ws.cell(row=row_num, column=2, value=key)
            ws.cell(row=row_num, column=3, value=value)
            row_num += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50


def _write_assay_conditions(wb: Workbook, assay_conditions: List[Any]) -> None:
    """
    Write AssayConditions sheet.

    When no conditions are provided, a blank template with default column
    headers is written so users can fill it in manually.

    :param wb: Workbook object
    :param assay_conditions: List of AssayCondition objects (may be empty)
    """
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("AssayConditions")

    # Add header comment
    ws.cell(row=1, column=1, value="# Assay Conditions - Per-well metadata")
    ws.cell(row=1, column=1).font = Font(italic=True, color="808080")

    # Collect condition keys from data, or fall back to defaults
    if assay_conditions:
        all_keys: set = set()
        for condition in assay_conditions:
            all_keys.update(condition.conditions.keys())
        condition_keys = sorted(all_keys)
    else:
        condition_keys = list(DEFAULT_CONDITION_KEYS)

    # Write headers
    headers = ["Plate", "Well"] + condition_keys
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write data rows
    for row_idx, condition in enumerate(assay_conditions, start=3):
        ws.cell(row=row_idx, column=1, value=condition.plate)
        ws.cell(row=row_idx, column=2, value=condition.well)

        for col_idx, key in enumerate(condition_keys, start=3):
            value = condition.conditions.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    for col_idx in range(3, 3 + len(condition_keys)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def _write_reference_sheet(wb: Workbook, sheet_name: str, data: Dict[str, Any]) -> None:
    """
    Write a reference sheet (sheets starting with _).

    :param wb: Workbook object
    :param sheet_name: Name of the reference sheet
    :param data: Dictionary of key-value pairs
    """
    # Ensure sheet name starts with underscore
    if not sheet_name.startswith('_'):
        sheet_name = f'_{sheet_name}'

    ws = wb.create_sheet(sheet_name)

    if not data:
        # Empty reference sheet
        ws.cell(row=1, column=1, value="# Empty reference sheet")
        return

    # Write headers
    headers = ["Key", "Value"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write data as key-value pairs
    for row_idx, (key, value) in enumerate(data.items(), start=2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
